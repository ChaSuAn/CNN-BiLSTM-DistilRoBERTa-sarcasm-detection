import random
import requests
import hashlib
import pyperclip
from openpyxl import load_workbook
data6 = load_workbook(r'C:\Users\86159\Documents\WeChat Files\wxid_lwvd11ru25ug22\FileStorage\File\2023-04\鹿晗.xls')  # 改三处
data = data6.active
url = "https://fanyi-api.baidu.com/api/trans/vip/translate"
appid = "20221210001492271"  # 这个双引号里替换成你的appid
password = "XmfkhoCiGZhcdMgUgY8c"  # 这个双引号里替换成你的密钥
ran = str(random.randint(1000000000,9999999999))

def get_text(raw):
    try:
        text = data.cell(raw, 6).value
    except:
        return
    return text
def get_sign(txt):
    sign = appid + txt + ran + password
    signMD5 = hashlib.md5(sign.encode('utf-8')).hexdigest()
    return signMD5

def get_result(signMD5,txt,place):
    r = requests.get(url+"?q="+txt+"&from=zh&to=en&appid="+appid+"&salt="+ran+"&sign="+signMD5)
    try:
        r = eval(r.text)
        print(r)
        r = r["trans_result"][0]
        data.cell(place,7).value = r["dst"]
    except Exception as e:
        pass

def main():
    count = 0
    for i in range(2, 20001):
        txt = get_text(i)
        signMD5 = get_sign(txt)
        get_result(signMD5, txt, i)
        count += 1
        print(count)
        if count % 50 == 0:
            data6.save(r'C:\Users\86159\Documents\WeChat Files\wxid_lwvd11ru25ug22\FileStorage\File\2023-04\鹿晗.xls')

if __name__ == "__main__":
    main()
    data6.save(r'C:\Users\86159\Documents\WeChat Files\wxid_lwvd11ru25ug22\FileStorage\File\2023-04\鹿晗.xls')
